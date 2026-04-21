from pathlib import Path
from openpyxl import load_workbook
import zipfile
import tempfile
import shutil
from xml.etree import ElementTree as ET

def ensure_folder(carpeta: Path):
  carpeta.mkdir(parents=True, exist_ok=True)

def delete_sheet(ruta_archivo, nombre_hoja):
  libro = load_workbook(ruta_archivo)
  if nombre_hoja in libro.sheetnames:
    del libro[nombre_hoja]
    libro.save(ruta_archivo)

def reorder_sheets(path_xlsx, sheet_order):
  """
  Reordena las hojas de un archivo Excel (.xlsx) según el orden dado en sheet_order.
  El proceso es rápido y no depende de openpyxl.
  """
  # Crear un archivo temporal
  temp_path = tempfile.mktemp(suffix=".xlsx")

  with zipfile.ZipFile(path_xlsx, 'r') as zin:
    with zipfile.ZipFile(temp_path, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
      for item in zin.infolist():
        data = zin.read(item.filename)

        # Solo modificar el archivo xl/workbook.xml
        if item.filename == "xl/workbook.xml":
          root = ET.fromstring(data)
          ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
          sheets = root.find('main:sheets', ns)
          if sheets is not None:
            # Crear un diccionario de las hojas existentes
            sheet_elements = {s.get('name'): s for s in sheets.findall('main:sheet', ns)}
            sheets.clear()  # limpiar todas
            # Reinsertar en el orden deseado
            for name in sheet_order:
              if name in sheet_elements:
                sheets.append(sheet_elements[name])
            data = ET.tostring(root, encoding='utf-8', xml_declaration=True)

        zout.writestr(item, data)

  # Reemplazar el original
  shutil.move(temp_path, path_xlsx)