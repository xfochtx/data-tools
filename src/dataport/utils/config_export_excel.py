from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle, Font
from openpyxl import load_workbook

# Método para colocar un formato específico a las fechas
def config_fecha(book, columna_fecha):
  # Establecer el formato de fecha en 'dd/mm/yyyy'
  date_style = NamedStyle(name='date_style')
  date_style.number_format = 'DD/MM/YYYY'
  date_style.font = Font(name='Calibri', size=11)  # Opcional: establecer la fuente y el tamaño de la fuente
  for sheetName in book.sheetnames:
    sheet = book[sheetName]
    # Obtener los nombres de las columnas desde la primera fila
    name_columns = [cell.value for cell in sheet[1]]
    pos_columna_fecha = name_columns.index(columna_fecha) + 1
    for cell in sheet[pos_columna_fecha]:
      cell.style = date_style

      # Método para configurar los anchos de las columnas


def config_width_col(book, width_col):
  for sheetName in book.sheetnames:
    sheet = book[sheetName]
    # Obtener los nombres de las columnas desde la primera fila
    name_columns = [cell.value for cell in sheet[1]]
    # Configurar el ancho de cada columna basado en los nombres de las columnas
    for idx, col_name in enumerate(name_columns, start=1):  # `start=1` para que el índice comience en 1
      if col_name in width_col:
        col_letter = sheet.cell(row=1, column=idx).column_letter  # Obtiene la letra identificadora de cada columna
        sheet.column_dimensions[col_letter].width = width_col[col_name]


# Método para configurar las alineaciones de las columnas
def config_align_col(book, align_col):
  for sheetName in book.sheetnames:
    sheet = book[sheetName]
    # Obtener los nombres de las columnas desde la primera fila
    name_columns = [cell.value for cell in sheet[1]]
    # Configurar la alineación de cada columna basado en los nombres de las columnas
    for idx, col_name in enumerate(name_columns, start=1):  # `start=1` para que el índice comience en 1
      if col_name in align_col:
        alignment = Alignment(horizontal=align_col[col_name], vertical='center')
        for cell in sheet[get_column_letter(idx)]:
          cell.alignment = alignment

def copy_format(ws, fila_template, fila_inicio, fila_fin):
  """Copia el number_format de la fila plantilla a todas las filas."""

  # Guardar formatos de la fila plantilla (solo una vez)
  formatos = [
    ws.cell(row=fila_template, column=col).number_format
    for col in range(1, ws.max_column + 1)
  ]

  # Aplicarlos a todas las filas
  for row in range(fila_inicio, fila_fin + 1):
    for col, fmt in enumerate(formatos, start=1):
      if fmt:
        ws.cell(row=row, column=col).number_format = fmt

def export_with_template(ruta_archivo, dataframe, nombre_hoja, ruta_plantilla, append=False):
  # 1) Abrir workbook correcto (write/append)
  if append and Path(ruta_archivo).exists():
    wb = load_workbook(ruta_archivo)
  else:
    wb = load_workbook(ruta_plantilla)

  # 2) Seleccionar hoja
  if nombre_hoja in wb.sheetnames:
    ws = wb[nombre_hoja]
  else:
    ws = wb[wb.sheetnames[0]]
    print(f"⚠️  La hoja '{nombre_hoja}' no existe en la plantilla. Se usará la hoja '{ws.title}'")

  # 3) Obtener tabla (se asume una tabla por hoja)
  table = ws.tables[list(ws.tables.keys())[0]]

  # 4) Fila donde empiezan los datos
  inicio_fila = 2

  # 5) Limpiar solo valores, (no formatos). Se optimizó para no borrar por celda, sino por fila
  if ws.max_row > 1:
    ws.delete_rows(2, ws.max_row)

  # 6) Convertir DataFrame a lista de listas (escritura mucho más rápida)
  data_as_list = dataframe.values.tolist()

  # 7) Escribir data masivamente (sin copiar estilo por celda). Se optimizó para no borrar por celda, sino por fila
  for row in data_as_list:
    ws.append(row)

  # 8) Actualizar rango de la tabla (Excel replicará estilos automáticamente)
  ultima_fila = inicio_fila + len(data_as_list) - 1

  # copiar formatos de número
  copy_format(ws, 2, inicio_fila, ultima_fila)

  last_col = get_column_letter(ws.max_column)
  table.ref = f"A1:{last_col}{ultima_fila}"

  ws.title = nombre_hoja
  wb.save(ruta_archivo)