"""Excel styling functions: date format, column width, alignment.

These operate on an openpyxl Workbook *after* writing data.

Usage:
    from dataport.excel import style

    style.config_date(wb, "Fecha")
    style.config_column_width(wb, {"Nombre": 25, "Valor": 12})
    style.config_alignment(wb, {"Nombre": "left", "Valor": "right"})
"""
from openpyxl.styles import Alignment, NamedStyle, Font
from openpyxl.utils import get_column_letter


def config_date(book, date_column: str):
  """Set 'DD/MM/YYYY' format on all cells in *date_column*."""
  date_style = NamedStyle(name="date_style")
  date_style.number_format = "DD/MM/YYYY"
  date_style.font = Font(name="Calibri", size=11)

  for sheet in book.worksheets:
    headers = [cell.value for cell in sheet[1]]
    try:
      col_idx = headers.index(date_column) + 1
    except ValueError:
      continue
    for cell in sheet[col_idx]:
      cell.style = date_style


def config_column_width(book, column_widths: dict[str, int]):
  """Set column widths by header name.

  Args:
    book: An openpyxl Workbook.
    column_widths: Dict mapping header name → width in characters.
  """
  for sheet in book.worksheets:
    headers = [cell.value for cell in sheet[1]]
    for idx, name in enumerate(headers, start=1):
      if name in column_widths:
        col_letter = get_column_letter(idx)
        sheet.column_dimensions[col_letter].width = column_widths[name]


def config_alignment(book, column_alignments: dict[str, str]):
  """Set horizontal alignment by header name.

  Args:
    book: An openpyxl Workbook.
    column_alignments: Dict mapping header name → "left", "center", "right".
  """
  for sheet in book.worksheets:
    headers = [cell.value for cell in sheet[1]]
    for idx, name in enumerate(headers, start=1):
      if name in column_alignments:
        alignment = Alignment(horizontal=column_alignments[name], vertical="center")
        col_letter = get_column_letter(idx)
        for cell in sheet[col_letter]:
          cell.alignment = alignment
