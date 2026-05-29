"""Internal helpers for Excel export: DataFrame-to-sheet, sheet reordering,
and sheet deletion.

These are private (`_`-prefixed) and should not be imported from outside
the excel package.
"""
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def write_df_to_sheet(ws, df: pd.DataFrame, index: bool = False):
  """Write a DataFrame into an openpyxl worksheet (header + data rows).

  Uses ws.append() with df.values.tolist() for speed.

  Note: converts pd.NA/NaT to None for openpyxl compatibility
  across all pandas versions.
  """
  df = df.where(df.notna(), None)
  if index:
    idx_name = df.index.name or ""
    ws.append([idx_name] + list(df.columns))
    for i, row in enumerate(df.values.tolist()):
      ws.append([df.index[i]] + row)
  else:
    ws.append(list(df.columns))
    for row in df.values.tolist():
      ws.append(row)


def reorder_sheets(wb, order: list[str]):
  """Reorder workbook sheets in-place to match *order*.

  Uses openpyxl's move_sheet() — no XML/ZIP manipulation needed.
  """
  for target_idx, name in enumerate(order):
    if name not in wb.sheetnames:
      continue
    current_idx = wb.sheetnames.index(name)
    if current_idx != target_idx:
      wb.move_sheet(name, offset=target_idx - current_idx)


def delete_sheet(path: Path, sheet_name: str):
  """Delete a sheet from an existing .xlsx file.

  Opens the workbook, deletes the named sheet, and saves.

  Args:
    path: Path to the .xlsx file.
    sheet_name: Name of the sheet to delete.
  """
  wb = load_workbook(path)
  if sheet_name in wb.sheetnames:
    del wb[sheet_name]
    wb.save(path)
