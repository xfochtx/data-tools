"""Template-based Excel export — preserves tables, formats, and styles.

Uses openpyxl load_workbook to open an existing .xlsx template,
find its Excel table, clear old data, write new data, and update
the table reference range.
"""
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def export_with_template(
  file_path: Path,
  df,
  sheet_name: str,
  template_path: Path,
  append: bool = False,
):
  """Write a DataFrame into an Excel file using a formatting template.

  The template must have an Excel table (structured reference) in the
  target sheet. Table formatting (colors, borders, fonts) is preserved.

  Args:
    file_path: Output .xlsx path.
    df: DataFrame to export.
    sheet_name: Target sheet in the template.
    template_path: Template .xlsx to load formatting from.
    append: If True, opens *file_path* instead of *template_path*
      (used internally by write_multi to append sheets).
  """
  if append and file_path.exists():
    wb = load_workbook(file_path)
  else:
    wb = load_workbook(template_path)

  # Select sheet
  if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
  else:
    ws = wb[wb.sheetnames[0]]
    print(f"⚠️  Sheet '{sheet_name}' not found in template. Using '{ws.title}'.")

  # Assume one table per sheet
  table = ws.tables[list(ws.tables.keys())[0]]

  # Capture number formats from the template data row (row 2) BEFORE clearing
  template_formats = [
    ws.cell(row=2, column=col).number_format
    for col in range(1, ws.max_column + 1)
  ]

  # Clear existing data (keep row 1 header)
  if ws.max_row > 1:
    ws.delete_rows(2, ws.max_row)

  # Write data (DataFrame → list of lists → ws.append)
  df = df.where(df.notna(), None)
  data = df.values.tolist()
  for row in data:
    ws.append(row)

  # Update table reference range
  last_row = 2 + len(data) - 1
  last_col = get_column_letter(ws.max_column)
  table.ref = f"A1:{last_col}{last_row}"

  # Apply captured formats to all new data rows
  _apply_formats(ws, template_formats, start_row=2, end_row=last_row)

  ws.title = sheet_name
  wb.save(file_path)


def _apply_formats(ws, formats: list[str], start_row: int, end_row: int):
  """Apply pre-captured number *formats* to all data rows.

  *formats* is a list where the index corresponds to the column number
  (1-based, e.g. formats[0] = column 1, formats[3] = column 4).
  """
  for row in range(start_row, end_row + 1):
    for col, fmt in enumerate(formats, start=1):
      if fmt and fmt != "General":
        ws.cell(row=row, column=col).number_format = fmt
