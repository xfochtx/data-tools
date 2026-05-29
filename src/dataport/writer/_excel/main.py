"""Excel export — single sheet and multi-sheet, with optional templates.

Usage:
    from dataport.excel import write, write_multi

    write(Path("out.xlsx"), df, sheet_name="Data")
    write_multi(Path("out.xlsx"), {"A": df1, "B": df2}, template_path="tpl.xlsx")
"""
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from ._helpers import write_df_to_sheet, reorder_sheets
from ._template import export_with_template
from ...utils.helpers import ensure_folder


def write(
  path: Path,
  df: pd.DataFrame,
  sheet_name: str = "Sheet1",
  template_path: Path | None = None,
  index: bool = False,
):
  """Export a DataFrame to an Excel file (single sheet).

  Uses openpyxl directly for speed and formatting control.

  When *template_path* is provided, loads the template and uses its
  formatting (tables, styles). Otherwise creates a new workbook.

  Args:
    path: Output .xlsx path.
    df: DataFrame to export.
    sheet_name: Sheet name in the workbook.
    template_path: Optional .xlsx template to use as formatting base.
    index: Whether to write row indices.
  """
  ensure_folder(path.parent)

  if template_path:
    export_with_template(path, df, sheet_name, template_path, append=False)
  else:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    write_df_to_sheet(ws, df, index=index)
    wb.save(path)


def write_multi(
  path: Path,
  dataframes: dict[str, pd.DataFrame],
  index: bool = False,
  template_path: Path | None = None,
):
  """Export multiple DataFrames to a single Excel file.

  Without a template, creates a single workbook, writes all sheets in
  memory, and saves once — no open/close overhead per sheet.

  With a template, loads the template and delegates each sheet to
  export_with_template() to preserve table references and formatting.

  Sheets are written in ascending row-count order (smallest first) for
  performance, then reordered back to the original dict order.

  Args:
    path: Output .xlsx path.
    dataframes: Dict mapping sheet names to DataFrames.
    index: Whether to write row indices.
    template_path: Optional .xlsx template for all sheets.
  """
  ensure_folder(path.parent)

  start_time = time.time()
  start_dt = datetime.now().strftime("%H:%M:%S")
  print(f"\n🕒 [Start] Multi-sheet export at {start_dt}")

  if path.exists():
    path.unlink()

  sorted_items = sorted(
    dataframes.items(),
    key=lambda item: len(item[1]) if hasattr(item[1], "__len__") else float("inf"),
  )
  original_order = list(dataframes.keys())

  if template_path:
    _write_multi_with_template(path, sorted_items, original_order, template_path, index)
  else:
    _write_multi_fresh(path, sorted_items, original_order, index)

  duration = time.time() - start_time
  print(f"✅ [Done] {duration:.2f}s")


def _write_multi_fresh(
  path: Path,
  sorted_items: list[tuple[str, pd.DataFrame]],
  original_order: list[str],
  index: bool,
):
  """Write multiple sheets into a brand-new workbook (no template)."""
  wb = Workbook()
  default = wb.active
  wb.remove(default)

  for sheet_name, df in sorted_items:
    sheet_start = time.time()
    print(f"  ➜ {sheet_name} ({len(df)} rows)... ", end="", flush=True)

    ws = wb.create_sheet(title=sheet_name)
    write_df_to_sheet(ws, df, index=index)

    print(f"✔️  {time.time() - sheet_start:.2f}s")

  reorder_sheets(wb, original_order)
  wb.save(path)


def _write_multi_with_template(
  path: Path,
  sorted_items: list[tuple[str, pd.DataFrame]],
  original_order: list[str],
  template_path: Path,
  index: bool,
):
  """Write multiple sheets reusing a template workbook.

  Each sheet delegates to export_with_template() to preserve table
  references and formatting. First call opens the template; subsequent
  calls append to the saved file.
  """
  for i, (sheet_name, df) in enumerate(sorted_items):
    sheet_start = time.time()
    print(f"  ➜ {sheet_name} ({len(df)} rows)... ", end="", flush=True)

    export_with_template(
      path, df, sheet_name, template_path,
      append=(i > 0),
    )

    print(f"✔️  {time.time() - sheet_start:.2f}s")
